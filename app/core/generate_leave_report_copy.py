"""
Generates a Leave Report .xlsx with fully dynamic leave type columns.
One merged header + two sub-columns (Available / Availed) per leave type.
"""

import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.schemas.leave.leave_report import LeaveReportResponse


# ── colour palette ────────────────────────────────────────────────────────────
HEADER_BG   = "1F4E79"
HEADER_FG   = "FFFFFF"
SUBHDR_BG   = "2E75B6"
SUBHDR_FG   = "FFFFFF"
ROW_ALT_BG  = "DEEAF1"
TOTAL_BG    = "F2F2F2"
AVAIL_BG    = "E2EFDA"
AVAIL_FG    = "375623"
AVAILED_BG  = "FCE4D6"
AVAILED_FG  = "843C0C"

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Arial")

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _cols_for(lt: dict) -> int:
    """EL types get 3 cols (available, availed, encashed), rest get 2."""
    return 3 if lt["code"].lower() in {"el_e", "el_ne"} else 2


def generate_leave_report_excel(report: LeaveReportResponse) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    # Fixed columns: EMP ID (A), EMP Name (B), Station (C)
    FIXED_COLS = 3

    # Dynamic leave type columns: 2 cols per leave type (available + availed)
    leave_types = report.leave_type_meta   # [{"code": "CL", "name": "Casual Leave"}, ...]
    total_cols  = FIXED_COLS + len(leave_types) * 2

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    for i in range(len(leave_types) * 2):
        col_letter = get_column_letter(FIXED_COLS + 1 + i)
        ws.column_dimensions[col_letter].width = 16

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 18

    last_col_letter = get_column_letter(total_cols)

    # ── Row 1 – Banner ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    banner = ws["A1"]
    banner_text = (
        f"Leave Report  |  From: {report.from_date.strftime('%d-%b-%Y')}"
        f"   To: {report.to_date.strftime('%d-%b-%Y')}"
    )
    if report.station:
        banner_text += f"   |  Station: {report.station}"
    banner.value     = banner_text
    banner.font      = _font(bold=True, color=HEADER_FG, size=12)
    banner.fill      = _fill(HEADER_BG)
    banner.alignment = _center()
    banner.border    = thin_border

    # ── Row 4 – Static column headers (EMP ID, Name, Station) ────────────────
    for col_idx, label in enumerate(["EMP ID", "EMP Name", "Station"], start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font      = _font(bold=True, color=HEADER_FG)
        cell.fill      = _fill(HEADER_BG)
        cell.alignment = _center()
        cell.border    = thin_border

    # Merge rows 4–5 for the 3 fixed columns
    for col_idx in range(1, FIXED_COLS + 1):
        ws.merge_cells(
            start_row=4, start_column=col_idx,
            end_row=5,   end_column=col_idx
        )

    # ── Row 4 – Dynamic leave type group headers (merged pairs) ───────────────
    for lt_idx, lt in enumerate(leave_types):
        start_col = FIXED_COLS + 1 + lt_idx * 2
        end_col   = start_col + 1
        start_ltr = get_column_letter(start_col)
        end_ltr   = get_column_letter(end_col)

        ws.merge_cells(f"{start_ltr}4:{end_ltr}4")
        cell = ws.cell(row=4, column=start_col, value=lt["name"])
        cell.font      = _font(bold=True, color=HEADER_FG)
        cell.fill      = _fill(SUBHDR_BG)
        cell.alignment = _center()
        cell.border    = thin_border
        ws.cell(row=4, column=end_col).border = thin_border

    # ── Row 5 – Sub-headers (Available / Availed per leave type) ─────────────
    for lt_idx in range(len(leave_types)):
        avail_col   = FIXED_COLS + 1 + lt_idx * 2
        availed_col = avail_col + 1

        for col, label, bg, fg in [
            (avail_col,   "Available\nLeave", AVAIL_BG,   AVAIL_FG),
            (availed_col, "Leave\nAvailed",   AVAILED_BG, AVAILED_FG),
        ]:
            cell = ws.cell(row=5, column=col, value=label)
            cell.font      = _font(bold=True, color=fg, size=9)
            cell.fill      = _fill(bg)
            cell.alignment = _center()
            cell.border    = thin_border

    # ── Row 6 – Note ──────────────────────────────────────────────────────────
    note_start = get_column_letter(FIXED_COLS + 1)
    ws.merge_cells(f"{note_start}6:{last_col_letter}6")
    note = ws[f"{note_start}6"]
    note.value     = "Available Leave = balance as on End Date  |  Leave Availed = days taken within the reporting period"
    note.font      = Font(italic=True, size=8, color="595959", name="Arial")
    note.alignment = Alignment(horizontal="center", vertical="center")
    note.border    = thin_border

    # ── Data rows (starting at row 7) ─────────────────────────────────────────
    data_start_row = 7

    # Build a code → list-index map for fast lookup
    lt_index = {lt["code"]: i for i, lt in enumerate(leave_types)}

    for idx, rec in enumerate(report.records):
        row    = data_start_row + idx
        is_alt = idx % 2 == 1
        row_bg = ROW_ALT_BG if is_alt else "FFFFFF"
        ws.row_dimensions[row].height = 18

        def write(col_idx, value, align=_center, bg=row_bg, bold=False, num_fmt=None):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font      = _font(bold=bold)
            cell.fill      = _fill(bg)
            cell.alignment = align()
            cell.border    = thin_border
            if num_fmt:
                cell.number_format = num_fmt

        write(1, rec.emp_id)
        write(2, rec.emp_name, align=_left)
        write(3, rec.station or "—", align=_left)

        # Build code → summary map for this employee
        lt_map = {lt.code: lt for lt in rec.leave_types}

        for lt_meta in leave_types:
            code      = lt_meta["code"]
            lt_i      = lt_index[code]
            avail_col = FIXED_COLS + 1 + lt_i * 2
            avd_col   = avail_col + 1
            summary   = lt_map.get(code)

            write(avail_col, float(summary.available) if summary else 0.0, bg=AVAIL_BG,   num_fmt="0.00")
            write(avd_col,   float(summary.availed)   if summary else 0.0, bg=AVAILED_BG, num_fmt="0.00")

    # ── Totals row ─────────────────────────────────────────────────────────────
    if report.records:
        total_row  = data_start_row + len(report.records)
        last_data  = total_row - 1
        ws.row_dimensions[total_row].height = 20

        ws.merge_cells(f"A{total_row}:C{total_row}")
        lbl = ws[f"A{total_row}"]
        lbl.value     = f"TOTAL  ({report.total_employees} employees)"
        lbl.font      = _font(bold=True)
        lbl.fill      = _fill(TOTAL_BG)
        lbl.alignment = _center()
        lbl.border    = thin_border

        for lt_i in range(len(leave_types)):
            for offset in range(2):
                col_idx    = FIXED_COLS + 1 + lt_i * 2 + offset
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(row=total_row, column=col_idx)
                cell.value         = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data})"
                cell.font          = _font(bold=True)
                cell.fill          = _fill(TOTAL_BG)
                cell.alignment     = _center()
                cell.border        = thin_border
                cell.number_format = "0.00"

    # ── Freeze header rows ─────────────────────────────────────────────────────
    ws.freeze_panes = ws[f"A{data_start_row}"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
